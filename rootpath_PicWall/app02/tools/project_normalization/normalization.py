import sys
import pandas as pd
import openpyxl

def usage():
    print("Usage: python3 normalization.py <input_file> <output_path>")
    sys.exit(1)

def read_384_wells(wells_file):
    """读取384孔板浓度信息,根据行（1,2,3）和列名（A,B,C）生成对应的字典，用于后续查找"""
    wells_dict = {}    
    wb = openpyxl.load_workbook(wells_file, read_only=True)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # 判断worksheet 是否为空，空就丢掉，给出warning。
        if ws.calculate_dimension() == 'A1:A1':
            print("Warning: \'{}\' is empty, please check! Sheet dimension: {}".format(sheetname, ws.calculate_dimension()))
            continue

        df = pd.DataFrame(ws.values).set_index(0)
        df.rename(columns=df.iloc[0])
        df = df[1:]
        df.index.names=[None]
        # df = df.T # 这个转置决定了 字典的key是column还是row。
        a_dict = df.to_dict()
        # print(a_dict)
        wells_dict.setdefault(sheetname, a_dict)
    wb.close() # close the workbook after reading
    return wells_dict

def write2excel(data_content, output_path):
    workbook = openpyxl.Workbook()  # create a new workbook
    worksheet = workbook.create_sheet(index=0, title="Sheet1")  # create a new worksheet and put it in first sheet.
    title_content = ["Source Plate Name", "Source Well", "Destination Plate Name", "Destination well", "Transfer Volume"]
    worksheet.append(title_content)

    for each in data_content:
        worksheet.append(each)
    workbook.save(output_path + '/workbook.xlsx')

def main():
    if len(sys.argv) != 3:
        usage()
    wells_file = sys.argv[1]   # input.xlsx
    output_path = sys.argv[2]  # /path/to/output/

    well_dict = read_384_wells(wells_file)
    # 转置后字典：
    # well_dict: {
    #   'Sheet1': {'rowname': {column1: cell, col2: cell,...},'B':{},'C':{},},
    #   'Sheet2': {'A': {1: 158, 2: 214,...},'B':{},'C':{},},
    # }
    # 非转置字典（第二版）
    # well_dict: {
    #   'Sheet1': {column1:{'rowname':cell,rowname2:cell},2:{},3:{},}
    #   'Sheet2': {1:{A:23,B:34},2:{A:34,B:38},},
    # }
    data_content = []
    for sheet, o in well_dict.items():
        for column, rows in o.items():
            for row, cell in rows.items():
                if str(cell) == "nan":
                    continue
                data_content.append([sheet, row+str(column), sheet.replace("S","D"), row+str(column), cell])
    # for sheet, o in well_dict.items():
    #     for rowname, columns in o.items():
    #         for column, cell in columns.items():
    #             if str(cell) == "nan":
    #                 continue
    #             data_content.append([sheet, rowname+str(column), sheet.replace("S","D"), rowname+str(column), cell])
    write2excel(data_content, output_path=output_path)


if __name__ == '__main__':
    main()