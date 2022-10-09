# -*- coding: utf-8 -*-
import sys
import openpyxl
import pandas as pd


def change_df(df):
    arr = df.values
    new_df = pd.DataFrame(arr[1:,1:], index=arr[1:,0], columns=arr[0,1:])
    new_df.index.name = arr[0,0]
    return new_df

def write2excel(data_content, output, times):
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0, title="Sheet1")
    title_content = ["Gene_ID","Source Plate Name","Source Well","Destination Plate Name","Destination well","Transfer Volume(nL)"]
    worksheet.append(title_content)
    for each in data_content:
        # print(each) # for debug
        worksheet.append(each)
    workbook.save(output + '/workbook{}.xlsx'.format(str(times)))

def read_excel(file_name, flag):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    df = pd.DataFrame(worksheet.values)
    df = change_df(df) # 
    if flag == 'to_dict':
        df_dict = df.to_dict('index') # index, series, records, split
        return df_dict
    return df

def deal_file1(file_name):
    df = read_excel(file_name, 'to_list')
    # sorted_df = df.sort_values('reads.errFree', ascending=True)
    return df

def deal_file2(file_name):
    df_dict = read_excel(file_name, 'to_dict')
    return df_dict


def generate_content(sorted_df, output, volume, times, df_dict=None):
    data_content = []
    # 生成384孔板编号
    used_well_list = [chr(j)+str(i) for i in range(1,24+1) for j in range(ord('A'),ord('P')+1)]
    # print(used_well_list)
    index = 0
    plate_index = 1
    for r in sorted_df.iterrows():
        if index >= 376: # 只取376孔板的数据
            index = 0
            plate_index += 1
        # fix a bug: 如果没有查到相应的孔板，则报错
        if df_dict == None:
            data_content.append([r[1].name, "PlateS1", r[1][0].split('.')[1], "PlateD"+str(plate_index), used_well_list[index], volume])
            data_content.append([r[1].name, "PlateS1", r[1][1].split('.')[1], "PlateD"+str(plate_index), used_well_list[index], volume])
            index += 1
        else:
            data_content.append([r[1].name, "PlateS1", df_dict[r[1].name]['MTP2.Primer.Pstn'].split('.')[1], "PlateD"+str(plate_index), used_well_list[index], volume])
            data_content.append([r[1].name, "PlateS1", df_dict[r[1].name]['MTP3.Primer.Pstn'].split('.')[1], "PlateD"+str(plate_index), used_well_list[index], volume])
            index += 1
    write2excel(data_content, output, times)


if __name__ == '__main__':
    volume = sys.argv[1] # 100nl
    file1 = sys.argv[2]
    file2 = sys.argv[3]
    output = sys.argv[4]
    sorted_df = deal_file1(file1)
    df_dict = deal_file2(file2)
    generate_content(sorted_df=sorted_df, output=output, volume=volume, times=1, df_dict=None)
    generate_content(sorted_df=sorted_df, output=output, volume=volume, times=2, df_dict=df_dict)
