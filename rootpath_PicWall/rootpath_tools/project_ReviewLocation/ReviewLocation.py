import sys, os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, GradientFill
from openpyxl.utils.dataframe import dataframe_to_rows

def usage():
    print("Usage: python3 ReviewLocation.py <input_file> <output_path>")
    sys.exit(1)


def read_clone_screening_table(input_table):
    data = pd.read_csv(input_table, sep="\t", dtype=str, encoding='utf-8')
    data = data.set_index('Sequence_number')
    # print(data)
    adict = data.to_dict('index')
    # print(adict)
    #{'1': {'plate_number': '1', 'hole_number': 'A1'}, '2': {'plate_number': '1', 'hole_number': 'B1'},
    return adict


def read_input_file(input_file, adict):
    """
    1) selected quirtirer
        1. #ident == 100,
        2. #ref_vs_al == TRUE,
        3. select lower #D40, if #ref is duplicated.
    """
    df = pd.read_csv(input_file, sep="\t", dtype=str, encoding='utf-8')

    df1 = df.loc[(df['#ident'] == '100.000') & (df['#ref_vs_aln']=='TRUE')]
    # print(df1
    # df['dup'] = "N"
    # df['dup'].loc[~df.sort_values('#D40 (smaller is better)').duplicated('#ref', keep='first')] ="Y"
    # print(df)
    # m =df.duplicated('#ref')
    # df['dup'] = df.apply(lambda x: "Y" if m[x[x==x].index]   else "N", axis=1)
    # df.to_csv("mytest.xls", sep="\t",index=False,header=True)
    # df['select'] = df.apply(lambda x: "Y" if x['#ident']=='100.000' and x['#ref_vs_aln']=='TRUE' else "N", axis=1)
    df1['#D40 (smaller is better)'] = df1['#D40 (smaller is better)'].astype(int)
    df2 = df1.sort_values('#D40 (smaller is better)', ascending=True).drop_duplicates('#ref', keep='first')
    df2['Well'] = df2['#qry'].map(lambda x:x.split("-")[1])
    df2['SourcePlate'] = df2['Well'].map(lambda x:adict[x]['plate_number'])
    df2['SourceWell'] = df2['Well'].map(lambda x:adict[x]['hole_number'])
    df2.Well = df2.Well.astype(int)
    df2 = df2.sort_values("Well")
    # print(df2)
    return df2

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Calibri", size=11,)


def set_font(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Calibri", size=11,)


def write2plate(df, output_path):
    workbook = openpyxl.Workbook()
    worksheet0 = workbook.create_sheet(index=0, title="Clone Screening input table")
    # write df to sheet0
    rows = dataframe_to_rows(df)
    print(rows)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            worksheet0.cell(row=r_idx, column=c_idx, value=value)
    
    # select columns for worksheet2
    # df = df.loc[df['selected'] == 'Y']
    alist = list(zip(df.SourcePlate,df.SourceWell))
    worksheet = workbook.create_sheet(index=1, title="Clone Screening Output table")

    start_row = 2 # used for draw 96 plates.
    start_col = 'H'
    plate=1
    for i in range(1, 5): # 1-4 plates
        # print("write plate {}".format(plate))
        header = ["-",1,2,3,4,5,6,7,8,9,10,11,12]
        # print(header)
        for index, each in enumerate(header):
            worksheet.cell(row=start_row, column=ord(start_col)-64+index, value=each) # column must be numeric not string.
        for x in range(ord('A'), ord('H')+1):
            absolute_position = [chr(x),]
            relative_position = [chr(x),]
            worksheet.cell(row=start_row+1, column=ord(start_col)-64, value=chr(x)) # write H column
            for col in range(1,13):
                abs_position = str(chr(ord(start_col)+col))+str(start_row +1)
                rel_position = str(chr(x))+str(col)
                if (str(plate), rel_position) in alist:
                    relative_position.append("√")
                    worksheet.cell(row=start_row +1, column=ord(start_col)-64+col, value="√")    
                else:
                    relative_position.append("-")
                absolute_position.append(abs_position)
            ### print(absolute_position)
            # print(relative_position)
            ## Set border line for each row.
            set_border(worksheet, 'I{}:T{}'.format(start_row+1, start_row+1))
            set_font(worksheet, 'H{}:T{}'.format(start_row, start_row+1))
            start_row += 1
        
        # write plate, merge plate cells. bold plate font.
        worksheet.merge_cells('F{}:G{}'.format(start_row-5,start_row-2))
        cell = worksheet.cell(row=start_row-5, column=ord(start_col)-64-2)
        cell.value = "plate{}".format(plate)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name="Calibri",size=36, color="000000",bold=True)

        plate +=1
        start_row += 2
    workbook.save(os.path.join(output_path, "plate.location.table.xlsx"))


def main():
    input_file = sys.argv[1]  # sum.txt, generated by script by Pro Zhou's pipeline
    output_path = sys.argv[2]

    base_dir = os.path.dirname(os.path.abspath(__file__))
    clone_screening_table = os.path.join(base_dir, "Clone_Screening_Input_table.xls")
    adict = read_clone_screening_table(input_table= clone_screening_table)
    df = read_input_file(input_file=input_file, adict=adict)
    write2plate(df, output_path)

if __name__ == '__main__':
    main()